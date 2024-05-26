import openpyxl

def getRowCount(file,sheetName):
   workbook= openpyxl.load_workbook(file)
   worksheet=workbook.get_sheet_by_name(sheetName)
   return worksheet.max_row


def getColumnCount(file,sheetName):
   workbook= openpyxl.load_workbook(file)
   worksheet=workbook.get_sheet_by_name(sheetName)
   return worksheet.max_column


def readData(file,sheetName,row,col):
   workbook= openpyxl.load_workbook(file)
   worksheet=workbook.get_sheet_by_name(sheetName)
   return worksheet.cell(row,col).value


def updateData(file,sheetName,row,col,data):
    workbook = openpyxl.load_workbook(file)
    worksheet=workbook.get_sheet_by_name(sheetName)
    worksheet.cell(row,col).value=data
    workbook.save(file)


